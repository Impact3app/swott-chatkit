"""
server.py — swott-chatkit
Architecture : openai-chatkit + openai-agents SDK (Python natif)
Persistence : Supabase (PostgreSQL)

Les agents sont définis dans agents_openai.py (copier-coller direct du "Get code" OpenAI).
"""

import os
import re
import json
import uuid
import base64
import asyncio
import io
from datetime import datetime, timezone
from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from chatkit.server import ChatKitServer, StreamingResult
from chatkit.store import Store, AttachmentStore, NotFoundError
from chatkit.types import (
    Page, Thread, ActiveStatus,
    AssistantMessageItem, AssistantMessageContent,
    AssistantMessageContentPartTextDelta,
    AssistantMessageContentPartDone,
    AssistantMessageContentPartAdded,
    ThreadItemAddedEvent, ThreadItemDoneEvent,
    FileAttachment,
)
from agents import Runner, RunConfig, trace
from supabase import create_client, Client

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Import des agents depuis le fichier OpenAI (copier-coller direct) ────────
from agents_openai import (
    agent_ifelse_json as classifier,
    cortex_routage,
    marcel_politique_achats as marcel,
    leonard_diag_orga as leonard,
    hector_n_gociation as hector,
    gustave_data_expert as gustave,
    eustache_ia as eustache,
    marguerite_ia as marguerite,
    luther_ia as luther,
    chan_ia as chan,
    savannah_ia as savannah,
    albert_ia as albert,
    catherine_ia as catherine,
    mich_le_ia as michele,
    achille_tco_decompo as achille,
    hypathie_juriste_contrats as hypathie,
    sherlock_sourcing_cadrage as sherlock_cadrage,
    sherlock_fast_json_ai as sherlock_fast,
    sherlock_deep,
    hercule_comparaison_d_offres as hercule,
    clint_ai as clint,
    barack_ai as barack,
    isaac_plan_d_action_orga as isaac,
    mazarin_diplomate as mazarin,
    sebus_excel_expert as sebus,
    franklin_cr as franklin,
    augustine_cdc as augustine,
    freya_benchmark_cadrage as freya_cadrage,
    freya_json as freya_fast,
    freya_deep,
    hilda_rfar as hilda,
    hermes_bilan_carbone as hermes,
    iris_processus_achats as iris,
    ariane_assistante_rh as ariane,
    cortex_core,
    jacques_strat_gie_portefeuilles as jacques_json,
    henry_leviers_achats as henry_json,
    jacques_ia,
    henry_ia,
)

# ── Supabase ─────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

WORKFLOW_ID = "wf_696b4c50579481908a889f44236f130108bc443970089c82"


# ═══════════════════════════════════════════════════════════════════════════════
# SUPABASE PERSISTENCE
# ═══════════════════════════════════════════════════════════════════════════════

def db_save_thread(thread_id: str, user_id: str, client_name: str = ""):
    try:
        supabase.table("threads").upsert({
            "thread_id": thread_id,
            "user_id": user_id,
            "client_name": client_name,
        }).execute()
    except Exception as e:
        print(f"[Supabase] db_save_thread error: {e}")

def db_append_message(thread_id: str, role: str, content: str, tokens_used: int = 0):
    try:
        msg_id = f"{role}_{uuid.uuid4().hex[:12]}"
        supabase.table("messages").insert({
            "id": msg_id,
            "thread_id": thread_id,
            "role": role,
            "content": content,
            "tokens_used": tokens_used,
        }).execute()
    except Exception as e:
        print(f"[Supabase] db_append_message error: {e}")

def db_get_history(thread_id: str) -> list:
    try:
        res = supabase.table("messages") \
            .select("role, content") \
            .eq("thread_id", thread_id) \
            .order("created_at") \
            .execute()
        return [{"role": r["role"], "content": r["content"]} for r in res.data]
    except Exception as e:
        print(f"[Supabase] db_get_history error: {e}")
        return []

def db_upload_file(thread_id: str, att_id: str, filename: str, raw_bytes: bytes, mime_type: str) -> str:
    """Upload un fichier dans Supabase Storage (bucket 'attachments').
    Retourne le storage_path ou '' en cas d'erreur."""
    try:
        storage_path = f"{thread_id}/{att_id}_{filename}"
        supabase.storage.from_("attachments").upload(
            path=storage_path,
            file=raw_bytes,
            file_options={"content-type": mime_type}
        )
        print(f"[Storage] Fichier uploadé: {storage_path}")
        return storage_path
    except Exception as e:
        print(f"[Storage] Erreur upload: {e}")
        return ""


def db_get_threads_for_user(user_id: str) -> list:
    try:
        res = supabase.table("threads") \
            .select("thread_id, title, created_at") \
            .eq("user_id", user_id) \
            .order("created_at", desc=True) \
            .limit(50) \
            .execute()
        threads = []
        for t in res.data:
            last = supabase.table("messages") \
                .select("content") \
                .eq("thread_id", t["thread_id"]) \
                .order("created_at", desc=True) \
                .limit(1) \
                .execute()
            last_msg = last.data[0]["content"] if last.data else None
            threads.append({
                "thread_id": t["thread_id"],
                "title": t.get("title"),
                "created_at": t["created_at"],
                "last_msg": last_msg,
            })
        return threads
    except Exception as e:
        print(f"[Supabase] db_get_threads_for_user error: {e}")
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# WORKFLOW
# ═══════════════════════════════════════════════════════════════════════════════

def _run_config():
    return RunConfig(trace_metadata={
        "__trace_source__": "swott-chatkit-python",
        "workflow_id": WORKFLOW_ID
    })

async def run_workflow(history: list) -> str:

    async def _run(agent, h=None):
        inp = list(h) if h is not None else list(history)
        r = await Runner.run(agent, input=inp, run_config=_run_config())
        return r

    with trace("Impact3_CorteX"):

        r_cls = await _run(classifier)
        parsed = r_cls.final_output
        category = parsed.category if hasattr(parsed, "category") else parsed.get("category", "")

        DIRECT = {
            "": cortex_routage,
            "WAIT_CONFIRMATION": cortex_routage,
            "DIAGNOSTIC_ORGANISATIONNEL": leonard,
            "PLAN_ACTION_OEP": isaac,
            "ANALYSE_DONNEES": gustave,
            "DECOMPOSITION_COUTS": achille,
            "JURIDIQUE_CONTRATS": hypathie,
            "COMPARAISON_OFFRES": hercule,
            "REDACTION_AO": clint,
            "MATURITE_ACHATS": barack,
            "CORTEX_CORE": cortex_core,
            "POLITIQUE_ACHATS": marcel,
            "PREPARATION_NEGOCIATION": hector,
            "EMAILS_COMMUNICATION": mazarin,
            "SEBUS_EXCEL": sebus,
            "COMPTE_RENDU_CR": franklin,
            "CAHIER_DES_CHARGES": augustine,
            "RFAR_LABEL_DIAGNOSTIC": hilda,
            "MESURE_IMPACT_CARBONE": hermes,
            "REDACTION_PROCESSUS_ACHATS": iris,
            "RH_ASSISTANCE": ariane,
        }

        if category in DIRECT:
            r = await _run(DIRECT[category])
            return r.final_output_as(str)

        elif category == "STRATEGIE_PORTEFEUILLE":
            r_j = await _run(jacques_json)
            smr_axis = r_j.final_output.smr_axis if hasattr(r_j.final_output, "smr_axis") else ""
            h2 = list(history) + [item.to_input_item() for item in r_j.new_items]
            SMR = {"SMR_E": eustache, "SMR_R": marguerite, "SMR_CSR": luther, "SMR_SH": chan}
            r = await _run(SMR.get(smr_axis, jacques_ia), h2)
            return r.final_output_as(str)

        elif category == "LEVIERS_OPTIMISATION_PROJET":
            r_h = await _run(henry_json)
            sml_axis = r_h.final_output.sml_axis if hasattr(r_h.final_output, "sml_axis") else ""
            h2 = list(history) + [item.to_input_item() for item in r_h.new_items]
            SML = {"SML_E": michele, "SML_R": albert, "SML_CSR": savannah, "SML_SH": catherine}
            r = await _run(SML.get(sml_axis, henry_ia), h2)
            return r.final_output_as(str)

        elif category == "SOURCING_MARCHE_FOURNISSEUR":
            r_cadrage = await _run(sherlock_cadrage)
            h2 = list(history) + [item.to_input_item() for item in r_cadrage.new_items]
            r_fast = await _run(sherlock_fast, h2)
            launch = r_fast.final_output.launch_deep if hasattr(r_fast.final_output, "launch_deep") else False
            if launch:
                r_deep = await _run(sherlock_deep, h2)
                return r_deep.final_output_as(str)
            return r_cadrage.final_output_as(str)

        elif category == "BENCHMARK_CONCURRENTIEL":
            r_cadrage = await _run(freya_cadrage)
            h2 = list(history) + [item.to_input_item() for item in r_cadrage.new_items]
            r_fast = await _run(freya_fast, h2)
            launch = r_fast.final_output.launch_deep if hasattr(r_fast.final_output, "launch_deep") else False
            if launch:
                r_deep = await _run(freya_deep, h2)
                return r_deep.final_output_as(str)
            return r_cadrage.final_output_as(str)

        else:
            r = await _run(cortex_routage)
            return r.final_output_as(str)


# ═══════════════════════════════════════════════════════════════════════════════
# EXTRACTION DE TEXTE (fichiers uploadés par l'utilisateur)
# ═══════════════════════════════════════════════════════════════════════════════

def extract_text(filename: str, content_bytes: bytes) -> str:
    ext = filename.lower().rsplit(".", 1)[-1]
    try:
        if ext == "pdf":
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(content_bytes))
            return "\n".join(p.extract_text() or "" for p in reader.pages).strip()
        elif ext in ("docx", "doc"):
            import docx
            doc = docx.Document(io.BytesIO(content_bytes))
            return "\n".join(p.text for p in doc.paragraphs).strip()
        elif ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
            lines = []
            for sheet in wb.worksheets:
                lines.append(f"[Feuille: {sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    row_str = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_str.strip():
                        lines.append(row_str)
            return "\n".join(lines).strip()
        elif ext == "txt":
            return content_bytes.decode("utf-8", errors="ignore")
        else:
            return f"[Fichier: {filename} — format non supporté]"
    except Exception as e:
        return f"[Erreur extraction {filename}: {e}]"


# ═══════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION DE FICHIERS (Excel depuis marqueurs agents)
# ═══════════════════════════════════════════════════════════════════════════════

def extract_file_markers(text: str) -> tuple:
    pattern = r'\[FILE:EXCEL\]\s*(.*?)\s*\[/FILE\]'
    matches = re.findall(pattern, text, re.DOTALL)
    files = []
    for raw_json in matches:
        try:
            data = json.loads(raw_json)
            filename = data.get("filename", "export.xlsx")
            b64 = generate_xlsx_b64(data)
            files.append({"type": "excel", "filename": filename, "b64": b64})
            print(f"[FILE] Excel généré: {filename}")
        except Exception as e:
            print(f"[FILE] Erreur génération Excel: {e}")
    clean_text = re.sub(pattern, '', text, flags=re.DOTALL).strip()
    return clean_text, files


def generate_xlsx_b64(data: dict) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    sheets = data.get("sheets", [])
    if not sheets:
        sheets = [{"name": "Données", "headers": data.get("headers", []), "rows": data.get("rows", [])}]
    for sheet_data in sheets:
        name = sheet_data.get("name", "Feuille")[:31]
        ws = wb.create_sheet(title=name)
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_data in rows:
                if col_idx - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[col_idx - 1])))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")


# ═══════════════════════════════════════════════════════════════════════════════
# IN-MEMORY STORES (ChatKit)
# ═══════════════════════════════════════════════════════════════════════════════

class InMemoryStore(Store):
    def __init__(self):
        self._threads = {}
        self._items = {}
        self._attachments = {}

    async def load_thread(self, thread_id, context):
        if thread_id not in self._threads:
            self._threads[thread_id] = Thread(
                id=thread_id, title=None,
                created_at=datetime.now(timezone.utc),
                status=ActiveStatus(),
                items=Page(data=[], has_more=False),
                metadata={}
            )
        return self._threads[thread_id]

    async def save_thread(self, thread, context):
        self._threads[thread.id] = thread

    async def load_thread_items(self, thread_id, after, limit, order, context):
        items = list(self._items.get(thread_id, {}).values())
        return Page(data=items, has_more=False)

    async def add_thread_item(self, thread_id, item, context):
        self._items.setdefault(thread_id, {})[item.id] = item

    async def save_item(self, thread_id, item, context):
        self._items.setdefault(thread_id, {})[item.id] = item

    async def load_item(self, thread_id, item_id, context):
        if thread_id not in self._items or item_id not in self._items[thread_id]:
            raise NotFoundError(f"Item {item_id} not found")
        return self._items[thread_id][item_id]

    async def delete_thread(self, thread_id, context):
        self._threads.pop(thread_id, None)
        self._items.pop(thread_id, None)

    async def delete_thread_item(self, thread_id, item_id, context):
        if thread_id in self._items:
            self._items[thread_id].pop(item_id, None)

    async def load_threads(self, limit, after, order, context):
        return Page(data=list(self._threads.values()), has_more=False)

    def generate_item_id(self, item_type, thread, context):
        return f"{item_type}_{uuid.uuid4().hex[:12]}"

    async def save_attachment(self, attachment, context):
        self._attachments[attachment.id] = attachment

    async def load_attachment(self, attachment_id, context):
        if attachment_id not in self._attachments:
            raise NotFoundError(f"Attachment {attachment_id} not found")
        return self._attachments[attachment_id]

    async def delete_attachment(self, attachment_id, context):
        self._attachments.pop(attachment_id, None)


class InMemoryAttachmentStore(AttachmentStore):
    def __init__(self, store: InMemoryStore):
        self._store = store
        self._bytes = {}
        self._texts = {}

    @property
    def _attachments(self):
        return self._store._attachments

    async def create_attachment(self, input, context):
        att_id = f"att_{uuid.uuid4().hex[:12]}"
        content = getattr(input, "content", None)
        thread_id = getattr(input, "thread_id", "")
        name = getattr(input, "name", "fichier")
        mime_type = getattr(input, "mime_type", "application/octet-stream")
        if content:
            raw = base64.b64decode(content) if isinstance(content, str) else content
            self._bytes[att_id] = raw
            self._texts[att_id] = extract_text(name, raw)
            print(f">>> Fichier reçu: {name} ({len(raw)} bytes)")
            # Upload dans Supabase Storage (bucket privé 'attachments')
            db_upload_file(thread_id, att_id, name, raw, mime_type)
        att = FileAttachment(
            id=att_id,
            thread_id=thread_id,
            created_at=datetime.now(timezone.utc),
            name=name,
            mime_type=mime_type,
            size=getattr(input, "size", len(self._bytes.get(att_id, b""))),
        )
        self._attachments[att_id] = att
        return att

    async def delete_attachment(self, attachment_id, context):
        self._store._attachments.pop(attachment_id, None)
        self._bytes.pop(attachment_id, None)
        self._texts.pop(attachment_id, None)

    def get_text(self, att_id):
        return self._texts.get(att_id, "")

    def get_attachment(self, att_id):
        return self._attachments.get(att_id)


# ═══════════════════════════════════════════════════════════════════════════════
# CHATKIT SERVER
# ═══════════════════════════════════════════════════════════════════════════════

class SwottChatKitServer(ChatKitServer):
    def __init__(self, data_store, attachment_store):
        super().__init__(data_store, attachment_store=attachment_store)
        self.att_store = attachment_store

    async def respond(self, thread, input, context):
        message_text = ""
        attachment_ids = []

        if input and hasattr(input, "content"):
            c = input.content
            if isinstance(c, str):
                message_text = c
            elif isinstance(c, list):
                for part in c:
                    if hasattr(part, "text"):
                        message_text += part.text
                    elif hasattr(part, "attachment_id"):
                        attachment_ids.append(part.attachment_id)

        if input and hasattr(input, "attachments") and input.attachments:
            for ref in input.attachments:
                aid = getattr(ref, "id", None) or getattr(ref, "attachment_id", None)
                if aid:
                    attachment_ids.append(aid)

        if attachment_ids:
            files_ctx = "\n\n---DOCUMENTS FOURNIS PAR L'UTILISATEUR---\n"
            for aid in attachment_ids:
                att = self.att_store.get_attachment(aid)
                text = self.att_store.get_text(aid)
                name = att.name if att else aid
                files_ctx += f"\n[Fichier: {name}]\n{text}\n"
            files_ctx += "---FIN DES DOCUMENTS---\n"
            message_text = message_text + files_ctx

        user_id = context.get("user_id", "anon") if isinstance(context, dict) else "anon"
        client_name = context.get("client_name", "") if isinstance(context, dict) else ""
        db_save_thread(thread.id, user_id, client_name)

        history = db_get_history(thread.id)
        history.append({"role": "user", "content": message_text or " "})

        print(f">>> [{thread.id}] user_id={user_id} msg={message_text[:80]}")

        # ── Appel workflow ──
        response_text = await run_workflow(history)
        response_text = re.sub(r'filecite\S+', '', response_text).strip()

        # ── Détecter et générer les fichiers ──
        response_text, generated_files = extract_file_markers(response_text)

        if generated_files:
            print(f"[FILE] {len(generated_files)} fichier(s) détecté(s) et généré(s)")

        # ── Persister dans Supabase (sans le base64) ──
        text_for_db = response_text
        if generated_files:
            filenames = ", ".join(f["filename"] for f in generated_files)
            text_for_db += f"\n\n📎 Fichier(s) généré(s) : {filenames}"
        tokens = len((message_text or " ").split()) + len(text_for_db.split())
        db_append_message(thread.id, "user", message_text or " ")
        db_append_message(thread.id, "assistant", text_for_db, tokens_used=tokens)

        print(f">>> Réponse: {response_text[:80]}")

        # ── Stream la réponse texte ──
        item_id = f"msg_{uuid.uuid4().hex[:12]}"
        now = datetime.now(timezone.utc)

        assistant_item = AssistantMessageItem(
            id=item_id, thread_id=thread.id, created_at=now,
            content=[AssistantMessageContent(type="output_text", text="", annotations=[])]
        )
        yield ThreadItemAddedEvent(item=assistant_item)
        yield AssistantMessageContentPartAdded(content_index=0, content={"type": "output_text", "text": ""})

        for i, word in enumerate(response_text.split(" ")):
            txt = word if i == 0 else " " + word
            yield AssistantMessageContentPartTextDelta(content_index=0, delta=txt)
            await asyncio.sleep(0.02)

        # ── Stream les fichiers générés ──
        for f in generated_files:
            file_html = (
                f'\n\n<!--SWOTT_FILE:{json.dumps({"filename": f["filename"], "type": f["type"], "b64": f["b64"]})}-->'
            )
            yield AssistantMessageContentPartTextDelta(content_index=0, delta=file_html)

        final_text = response_text
        if generated_files:
            for f in generated_files:
                final_text += f'\n\n<!--SWOTT_FILE:{json.dumps({"filename": f["filename"], "type": f["type"], "b64": f["b64"]})}-->'

        yield AssistantMessageContentPartDone(
            content_index=0,
            content={"type": "output_text", "text": final_text}
        )
        yield ThreadItemDoneEvent(item=AssistantMessageItem(
            id=item_id, thread_id=thread.id, created_at=now,
            content=[AssistantMessageContent(type="output_text", text=final_text, annotations=[])]
        ))


# ═══════════════════════════════════════════════════════════════════════════════
# FASTAPI
# ═══════════════════════════════════════════════════════════════════════════════

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

data_store = InMemoryStore()
att_store  = InMemoryAttachmentStore(data_store)
server     = SwottChatKitServer(data_store, att_store)


@app.post("/chatkit")
async def chatkit_endpoint(request: Request):
    body    = await request.body()
    user_id = request.headers.get("x-user-id") or request.query_params.get("user_id", "anon")
    client_name = request.headers.get("x-client-name") or request.query_params.get("client_name", "")
    result  = await server.process(body, {"user_id": user_id, "client_name": client_name})
    if isinstance(result, StreamingResult):
        return StreamingResponse(result, media_type="text/event-stream")
    return Response(content=result.json, media_type="application/json")


@app.get("/history/{thread_id}")
async def get_history(thread_id: str):
    messages = db_get_history(thread_id)
    return {"thread_id": thread_id, "messages": messages}


@app.get("/threads")
async def get_threads(user_id: str = "anon"):
    threads = db_get_threads_for_user(user_id)
    return {"user_id": user_id, "threads": threads}


@app.get("/download")
async def download_file(path: str):
    """Génère une URL signée temporaire (1h) pour télécharger un fichier depuis Supabase Storage."""
    try:
        result = supabase.storage.from_("attachments").create_signed_url(path, 3600)
        if result and result.get("signedURL"):
            return {"url": result["signedURL"]}
        return {"error": "Fichier non trouvé"}
    except Exception as e:
        return {"error": str(e)}


@app.get("/files/{thread_id}")
async def list_files(thread_id: str):
    """Liste les fichiers uploadés dans un thread."""
    try:
        files = supabase.storage.from_("attachments").list(thread_id)
        return {"thread_id": thread_id, "files": [
            {"name": f["name"], "size": f.get("metadata", {}).get("size", 0),
             "path": f"{thread_id}/{f['name']}"}
            for f in files if f.get("name")
        ]}
    except Exception as e:
        return {"thread_id": thread_id, "files": [], "error": str(e)}


@app.get("/stats")
async def get_stats(user_id: str = None):
    try:
        query = supabase.table("threads") \
            .select("user_id, client_name, thread_id") \
            .execute()
        threads = query.data
        stats = {}
        for t in threads:
            key = t["user_id"]
            if key not in stats:
                stats[key] = {
                    "user_id": t["user_id"],
                    "client_name": t.get("client_name") or "",
                    "nb_conversations": 0,
                    "nb_messages": 0,
                    "total_tokens": 0
                }
            stats[key]["nb_conversations"] += 1
            if not stats[key]["client_name"] and t.get("client_name"):
                stats[key]["client_name"] = t.get("client_name")
            msgs = supabase.table("messages") \
                .select("id, tokens_used") \
                .eq("thread_id", t["thread_id"]) \
                .eq("role", "assistant") \
                .execute()
            stats[key]["nb_messages"] += len(msgs.data)
            stats[key]["total_tokens"] += sum(m["tokens_used"] or 0 for m in msgs.data)
        for r in stats.values():
            if not r["client_name"]:
                r["client_name"] = "—"
        result = list(stats.values())
        if user_id:
            result = [r for r in result if r["user_id"] == user_id]
        return {"stats": sorted(result, key=lambda x: x["total_tokens"], reverse=True)}
    except Exception as e:
        return {"error": str(e)}

@app.get("/test-storage")
async def test_storage():
    try:
        test_data = b"test swott storage"
        result = supabase.storage.from_("attachments").upload(
            path="test/test.txt",
            file=test_data,
            file_options={"content-type": "text/plain"}
        )
        return {"status": "ok", "result": str(result)}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

@app.get("/")
async def healthcheck():
    return {"status": "ok", "service": "swott-chatkit", "mode": "agents-python-native+supabase"}
